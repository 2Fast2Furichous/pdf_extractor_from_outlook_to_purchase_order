"""
PDF Extractor from Outlook Emails - Eel Version (FIXED FOR MULTIPLE RUNS)
Modern web-based UI with EXACT original functionality preserved
Fixed: COM connection handling for multiple extractions
"""

import eel
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime
import os
import csv
import tempfile
import re
import sys
import json
import hashlib
import threading
import shutil

# Windows imports with proper handling
try:
    import win32timezone
except ImportError:
    pass  # Continue without it

import win32com.client
import pythoncom
import pywintypes
import pdfplumber
import pandas as pd

# Initialize Eel with the web folder
eel.init('web')

# Global variable for progress updates
current_status = {"message": "", "progress": 0}

class PDFExtractor:
    def __init__(self):
        self.settings_file = os.path.join(os.path.expanduser("~"), ".pdf_extractor_settings.json")
        # Don't store outlook - create fresh connection each time
        
    def connect_outlook(self):
        """Connect to Outlook - creates fresh connection each time"""
        try:
            # Create a new connection each time
            outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
            return outlook
        except Exception as e:
            update_progress(f"Error connecting to Outlook: {e}")
            raise
    
    def find_folder(self, outlook, email_addr, folder_text):
        """Find folder in Outlook - EXACT ORIGINAL LOGIC"""
        try:
            # Try to get account by email
            for account in outlook.Folders:
                if email_addr.lower() in account.Name.lower():
                    return self.search_subfolder(account, folder_text)
            
            # If not found, search all folders
            for account in outlook.Folders:
                folder = self.search_subfolder(account, folder_text)
                if folder:
                    return folder
            
            return None
        except Exception as e:
            update_progress(f"Error finding folder: {e}")
            return None
    
    def search_subfolder(self, parent_folder, search_text):
        """Recursively search for subfolder - EXACT ORIGINAL LOGIC"""
        try:
            if search_text.lower() in parent_folder.Name.lower():
                return parent_folder
            
            for subfolder in parent_folder.Folders:
                result = self.search_subfolder(subfolder, search_text)
                if result:
                    return result
            
            return None
        except:
            return None
    
    def filter_emails(self, folder, subject_text, start_date, end_date):
        """Filter emails by criteria - EXACT ORIGINAL LOGIC"""
        matching_emails = []
        
        try:
            items = folder.Items
            items.Sort("[ReceivedTime]", True)  # Sort by received time, descending
            
            for item in items:
                try:
                    # Check if it's a mail item
                    if item.Class != 43:  # 43 = olMail
                        continue
                    
                    # Check subject
                    if subject_text and subject_text.lower() not in item.Subject.lower():
                        continue
                    
                    # Check date range
                    try:
                        received_date = item.ReceivedTime
                        # Compare just the date parts (year, month, day) to avoid type issues
                        received_date_only = datetime(received_date.year, received_date.month, received_date.day)
                        
                        # Check start date
                        if start_date and received_date_only < start_date:
                            continue
                        
                        # Check end date
                        if end_date and received_date_only > end_date:
                            continue
                    except Exception as date_err:
                        # If date conversion fails, log and skip date check for this email
                        update_progress(f"    Warning: Date conversion failed: {date_err}")
                        pass
                    
                    matching_emails.append(item)
                except:
                    continue
        
        except Exception as e:
            update_progress(f"Error filtering emails: {e}")
        
        return matching_emails
    
    def parse_pdf(self, pdf_path, pdf_name):
        """Parse PDF using pdfplumber - EXACT ORIGINAL LOGIC"""
        data = []
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                # Extract text from all pages
                full_text = ""
                for page in pdf.pages:
                    full_text += page.extract_text() + "\n"
                
                # Check for tables
                has_tables = False
                for page in pdf.pages:
                    if page.extract_tables():
                        has_tables = True
                        break
                
                update_progress(f"    Tables detected: {has_tables}")
                
                # Use table-based parsing if available
                if has_tables:
                    data = self.parse_pdf_tables(pdf, pdf_name)
                else:
                    data = self.parse_pdf_text(pdf, pdf_name)
                
                update_progress(f"    Extracted {len(data)} line items")
        
        except Exception as e:
            update_progress(f"    ERROR parsing PDF: {e}")
        
        return data
    
    def parse_pdf_tables(self, pdf, pdf_name):
        """Parse PDF using table extraction - handles multi-page line items - EXACT ORIGINAL LOGIC"""
        data = []
        
        try:
            # Extract order info from text and coordinates
            full_text = ""
            for page in pdf.pages:
                full_text += page.extract_text() + "\n"
            
            order_number = self.extract_order_number(full_text)
            order_date = self.extract_order_date(full_text)
            
            # Use first page for coordinate-based extraction of addresses
            first_page = pdf.pages[0] if pdf.pages else None
            ship_to = self.extract_ship_to_coordinates(first_page) if first_page else ""
            ordering_office = self.extract_ordering_office_coordinates(first_page) if first_page else ""
            
            # Track column mappings for continuation pages
            col_line = 0  # Line is always first
            col_part = None
            col_date = None
            col_qty = None
            col_unit_price = None
            col_amount = None
            found_header = False
            
            # Look for line items table - check all pages
            for page_num, page in enumerate(pdf.pages):
                tables = page.extract_tables()
                
                # Check all tables on this page
                for table_idx, table in enumerate(tables):
                    if not table or len(table) < 1:
                        continue
                    
                    # Try to detect if this is a line items table
                    first_row = table[0]
                    first_cell = str(first_row[0]).strip() if first_row and first_row[0] else ""
                    
                    # Check if first row is a header (contains "Line" and "Part")
                    is_header = False
                    if len(table) >= 2:
                        header = ' '.join([str(cell) for cell in first_row if cell])
                        if 'Line' in header and 'Part' in header:
                            is_header = True
                            found_header = True
                            
                            # Parse column indices from header
                            col_line = 0  # Line is always first
                            for idx, cell in enumerate(first_row):
                                if cell:
                                    cell_lower = str(cell).lower()
                                    if 'part' in cell_lower:
                                        col_part = idx
                                    elif 'delivery' in cell_lower or 'date' in cell_lower:
                                        col_date = idx
                                    elif 'quantity' in cell_lower:
                                        col_qty = idx
                                    elif 'unit price' in cell_lower or 'price' in cell_lower:
                                        col_unit_price = idx
                                    elif 'amount' in cell_lower or 'total' in cell_lower:
                                        col_amount = idx
                    
                    # Check if this is a continuation table (starts with line number like "4.1")
                    is_continuation = False
                    if not is_header and found_header and re.match(r'^\d+\.\d+$', first_cell):
                        is_continuation = True
                        
                        # For continuation tables, detect columns from first data row
                        # because page breaks can shift column positions
                        if len(first_row) > 1:
                            # Reset price/amount for this table's detection
                            table_unit_price = None
                            table_amount = None
                            
                            # Examine first row to find column positions
                            for idx, cell in enumerate(first_row):
                                cell_str = str(cell).strip() if cell else ""
                                # Date column - look for date pattern
                                if re.search(r'\d{1,2}-[A-Za-z]{3}-\d{4}', cell_str):
                                    col_date = idx
                                # Quantity - numeric value (but not price with decimal)
                                elif cell_str.isdigit():
                                    col_qty = idx
                                # UOM - typically "Each"
                                elif cell_str.lower() == 'each':
                                    pass  # We don't need UOM column index
                                # Unit Price - has decimal point, larger than quantity
                                elif re.match(r'^\d+\.\d{2,}$', cell_str):
                                    if not table_unit_price:
                                        col_unit_price = idx
                                        table_unit_price = idx
                                    else:
                                        col_amount = idx
                                        table_amount = idx
                    
                    # Skip if this is neither a header table nor a continuation
                    if not is_header and not is_continuation:
                        continue
                    
                    # Process data rows (skip header if present)
                    start_row = 1 if is_header else 0
                    for row in table[start_row:]:
                        if not row:
                            continue
                        
                        line_num = str(row[col_line]).strip() if row[col_line] else ""
                        
                        # Check if valid line number (e.g., "1.1", "2.1", "4.1")
                        if re.match(r'^\d+\.\d+$', line_num):
                            part_num = str(row[col_part]).strip() if col_part and len(row) > col_part and row[col_part] else ""
                            delivery_date = str(row[col_date]).strip() if col_date and len(row) > col_date and row[col_date] else ""
                            quantity = str(row[col_qty]).strip() if col_qty and len(row) > col_qty and row[col_qty] else ""
                            unit_price = str(row[col_unit_price]).strip() if col_unit_price and len(row) > col_unit_price and row[col_unit_price] else ""
                            amount = str(row[col_amount]).strip() if col_amount and len(row) > col_amount and row[col_amount] else ""
                            
                            # Clean part number (remove /REV: and newlines)
                            if '/' in part_num:
                                part_num = part_num.split('/')[0].strip()
                            part_num = part_num.replace('\n', ' ').strip()
                            
                            # Format delivery date to YYYYMMDD
                            delivery_date = self.format_date_to_yyyymmdd(delivery_date)
                            
                            data.append({
                                'pdf_file': pdf_name,
                                'order_number': order_number,
                                'order_date': order_date,
                                'line': line_num,
                                'part_number': part_num,
                                'quantity': quantity,
                                'unit_price': unit_price,
                                'amount': amount,
                                'delivery_date': delivery_date,
                                'ship_to': ship_to,
                                'ordering_office': ordering_office
                            })
        
        except Exception as e:
            update_progress(f"    Error in table parsing: {e}")
        
        return data
    
    def parse_pdf_text(self, pdf, pdf_name):
        """Parse PDF using text extraction (for vertical format) - EXACT ORIGINAL LOGIC"""
        data = []
        
        try:
            # Extract full text
            text = ""
            for page in pdf.pages:
                text += page.extract_text() + "\n"
            
            lines = text.split('\n')
            
            order_number = self.extract_order_number(text)
            order_date = self.extract_order_date(text)
            
            # Use coordinate-based extraction for addresses
            first_page = pdf.pages[0] if pdf.pages else None
            ship_to = self.extract_ship_to_coordinates(first_page) if first_page else ""
            ordering_office = self.extract_ordering_office_coordinates(first_page) if first_page else ""
            
            # Find line items section
            in_line_items = False
            i = 0
            
            while i < len(lines):
                line = lines[i].strip()
                
                # Detect start of line items
                if not in_line_items and 'line' in line.lower() and len(line) < 50:
                    in_line_items = True
                    i += 1
                    continue
                
                if in_line_items:
                    # Check for line number (e.g., "1.1", "2.1")
                    if re.match(r'^\d+\.\d+$', line):
                        line_num = line
                        part_num = ""
                        delivery_date = ""
                        quantity = ""
                        unit_price = ""
                        amount = ""
                        
                        # Get part number from next line
                        if i + 1 < len(lines):
                            part_line = lines[i + 1].strip()
                            if '/' in part_line:
                                part_num = part_line.split('/')[0].strip()
                            else:
                                part_num = part_line
                        
                        # Look for delivery date, quantity, price, and amount in next 10 lines
                        for j in range(i + 1, min(i + 11, len(lines))):
                            check_line = lines[j].strip()
                            
                            # Check for date pattern
                            if not delivery_date:
                                date_match = re.search(r'\d{1,2}-[A-Za-z]{3}-\d{4}', check_line)
                                if date_match:
                                    delivery_date = date_match.group()
                            
                            # Check for quantity (numeric, after date)
                            if delivery_date and not quantity:
                                if check_line.isdigit() and len(check_line) < 10:
                                    quantity = check_line
                            
                            # Check for price patterns (e.g., "$12.34" or "12.34")
                            if not unit_price:
                                price_match = re.search(r'\$?\d+\.\d{2}', check_line)
                                if price_match and not amount:
                                    unit_price = price_match.group()
                                elif price_match:
                                    amount = price_match.group()
                        
                        # Format delivery date to YYYYMMDD
                        delivery_date = self.format_date_to_yyyymmdd(delivery_date)
                        
                        if part_num:
                            data.append({
                                'pdf_file': pdf_name,
                                'order_number': order_number,
                                'order_date': order_date,
                                'line': line_num,
                                'part_number': part_num,
                                'quantity': quantity,
                                'unit_price': unit_price,
                                'amount': amount,
                                'delivery_date': delivery_date,
                                'ship_to': ship_to,
                                'ordering_office': ordering_office
                            })
                
                i += 1
        
        except Exception as e:
            update_progress(f"    Error in text parsing: {e}")
        
        return data
    
    def extract_order_number(self, text):
        """Extract 10-digit order number - EXACT ORIGINAL LOGIC"""
        match = re.search(r'\b(\d{10})\b', text)
        return match.group(1) if match else ""
    
    def extract_order_date(self, text):
        """Extract order date and format as YYYYMMDD - EXACT ORIGINAL LOGIC"""
        match = re.search(r'\d{1,2}-[A-Za-z]{3}-\d{4}', text)
        if match:
            try:
                # Parse the date (format like "1-Jan-2024")
                date_obj = datetime.strptime(match.group(), "%d-%b-%Y")
                # Return in YYYYMMDD format
                return date_obj.strftime("%Y%m%d")
            except ValueError:
                # If parsing fails, return the original
                return match.group()
        return ""
    
    def format_date_to_yyyymmdd(self, date_str):
        """Convert date string to YYYYMMDD format - EXACT ORIGINAL LOGIC"""
        if not date_str:
            return ""
        try:
            # Parse the date (format like "1-Jan-2024" or "01-Jan-2024")
            date_obj = datetime.strptime(date_str.strip(), "%d-%b-%Y")
            # Return in YYYYMMDD format
            return date_obj.strftime("%Y%m%d")
        except ValueError:
            # If parsing fails, return the original
            return date_str
    
    def extract_ship_to_coordinates(self, page):
        """Extract ship to address using pdfplumber coordinate-based extraction - EXACT ORIGINAL LOGIC"""
        try:
            words = page.extract_words()
            
            # Find "Ship To Address" label position
            ship_label_y = None
            for word in words:
                if 'Ship' in word['text'] and word.get('top', 0) > 100:  # Skip header
                    # Look for words near this one to confirm "Ship To Address"
                    ship_label_y = word['top']
                    break
            
            if not ship_label_y:
                return ""
            
            # Find "Payment Terms" label to know where to stop
            payment_terms_y = None
            for word in words:
                if 'Payment' in word['text'] and word['top'] > ship_label_y:
                    payment_terms_y = word['top']
                    break
            
            # Extract words in LEFT column (x < 300) between ship_label and payment_terms
            ship_words = []
            for word in words:
                word_y = word['top']
                word_x = word['x0']
                
                # Must be below the "Ship To Address" label and above "Payment Terms"
                if word_y > ship_label_y + 10:  # Start below the label
                    if payment_terms_y and word_y >= payment_terms_y:
                        break
                    # Left column only (ship-to, not invoice)
                    if word_x < 300:  # Adjust threshold as needed
                        ship_words.append((word_y, word_x, word['text']))
            
            # Sort by y-coordinate first, then x-coordinate (to get correct reading order)
            ship_words.sort(key=lambda w: (w[0], w[1]))
            
            # Group words by line (same Y coordinate within tolerance)
            lines = []
            current_line = []
            last_y = None
            
            for word_y, word_x, text in ship_words:
                if last_y is None or abs(word_y - last_y) < 2:  # Same line
                    current_line.append(text)
                    last_y = word_y
                else:  # New line
                    if current_line:
                        lines.append(' '.join(current_line))
                    current_line = [text]
                    last_y = word_y
            
            if current_line:  # Don't forget the last line
                lines.append(' '.join(current_line))
            
            # Join all lines and clean up
            address = ', '.join(lines[:5])  # Limit to first 5 lines
            address = address.replace(', ,', ',').strip(' ,')
            return address[:300] if address else ""
        
        except Exception as e:
            update_progress(f"Error in coordinate extraction for ship_to: {e}")
            return ""
    
    def extract_ordering_office_coordinates(self, page):
        """Extract ordering office using pdfplumber coordinate-based extraction - EXACT ORIGINAL LOGIC"""
        try:
            words = page.extract_words()
            
            # Find "Ordering Office" label position
            ordering_label_y = None
            for word in words:
                if 'Ordering' in word['text']:
                    ordering_label_y = word['top']
                    break
            
            if not ordering_label_y:
                return ""
            
            # Find "Supplier Contact" or "Buyer" to know where to stop
            stop_y = None
            for word in words:
                if word['top'] > ordering_label_y:
                    if 'Supplier' in word['text'] and 'Contact' in [w['text'] for w in words if abs(w['top'] - word['top']) < 5]:
                        stop_y = word['top']
                        break
                    if 'Buyer' in word['text']:
                        stop_y = word['top']
                        break
            
            # Extract words in RIGHT column (x > 300) between ordering_label and stop
            ordering_words = []
            for word in words:
                word_y = word['top']
                word_x = word['x0']
                
                # Must be below the "Ordering Office" label
                if word_y > ordering_label_y + 10:
                    if stop_y and word_y >= stop_y:
                        break
                    # Right column only (x > 300)
                    if word_x > 300:
                        ordering_words.append((word_y, word_x, word['text']))
            
            # Sort by y-coordinate first, then x-coordinate (to get correct reading order)
            ordering_words.sort(key=lambda w: (w[0], w[1]))
            
            # Group words by line (same Y coordinate within tolerance)
            lines = []
            current_line = []
            last_y = None
            
            for word_y, word_x, text in ordering_words:
                if last_y is None or abs(word_y - last_y) < 2:  # Same line
                    current_line.append(text)
                    last_y = word_y
                else:  # New line
                    if current_line:
                        lines.append(' '.join(current_line))
                    current_line = [text]
                    last_y = word_y
            
            if current_line:  # Don't forget the last line
                lines.append(' '.join(current_line))
            
            # Join all lines and clean up
            office = ', '.join(lines[:6])  # Limit to first 6 lines
            office = office.replace(', ,', ',').strip(' ,')
            return office[:300] if office else ""
        
        except Exception as e:
            update_progress(f"Error in coordinate extraction for ordering_office: {e}")
            return ""
    
    def write_output(self, output_path, data):
        """Write data to Excel or CSV file (append mode if exists) - EXACT ORIGINAL LOGIC WITH RETRY"""
        try:
            output_path = os.path.abspath(os.path.normpath(output_path))

            if not data:
                update_progress("Warning: No data to write")
                return
            
            fieldnames = ['pdf_file', 'order_number', 'order_date', 'line',
                         'part_number', 'quantity', 'unit_price', 'amount',
                         'delivery_date', 'ship_to', 'ordering_office']
            
            # Create DataFrame from new data
            new_df = pd.DataFrame(data, columns=fieldnames)
            
            # Rename columns to Title Case for better readability
            new_df.columns = ['PDF File', 'Order Number', 'Order Date', 'Line',
                             'Part Number', 'Quantity', 'Unit Price', 'Amount',
                             'Delivery Date', 'Ship To', 'Ordering Office']
            
            # Convert numeric columns to proper numeric types
            numeric_columns = ['Quantity', 'Unit Price', 'Amount']
            for col in numeric_columns:
                if col in new_df.columns:
                    # Remove commas and convert to numeric
                    new_df[col] = new_df[col].astype(str).str.replace(',', '')
                    new_df[col] = pd.to_numeric(new_df[col], errors='coerce')
            
            # Convert identifier columns to numeric to prevent Excel warnings
            # Order Number: convert to integer
            if 'Order Number' in new_df.columns:
                new_df['Order Number'] = pd.to_numeric(new_df['Order Number'], errors='coerce').astype('Int64')
            
            # Line: keep as string
            if 'Line' in new_df.columns:
                new_df['Line'] = new_df['Line'].astype(str)
            
            # Check if file exists and append if it does
            if os.path.exists(output_path):
                update_progress(f"File exists, attempting to append to: {output_path}")
                
                # Read existing file
                if output_path.endswith('.xlsx'):
                    existing_df = pd.read_excel(output_path)
                else:
                    existing_df = pd.read_csv(output_path)
                
                update_progress(f"Successfully read existing file with {len(existing_df)} rows")
                
                # CRITICAL: Convert existing data to match new data types before combining
                # This prevents duplicate detection failures due to type mismatches
                # Convert Order Number to Int64 (must match new_df)
                if 'Order Number' in existing_df.columns:
                    existing_df['Order Number'] = pd.to_numeric(existing_df['Order Number'], errors='coerce').astype('Int64')
                
                # Convert Line to string (must match new_df)
                if 'Line' in existing_df.columns:
                    existing_df['Line'] = existing_df['Line'].astype(str)
                
                # Convert numeric columns to numeric types (must match new_df)
                for col in ['Quantity', 'Unit Price', 'Amount']:
                    if col in existing_df.columns:
                        existing_df[col] = existing_df[col].astype(str).str.replace(',', '')
                        existing_df[col] = pd.to_numeric(existing_df[col], errors='coerce')
                
                # Combine old and new data
                combined_df = pd.concat([existing_df, new_df], ignore_index=True)
                
                # Remove duplicates based on key columns (PDF file, order number, and line number)
                # This prevents removing legitimate duplicate part numbers across different orders
                before_dedup = len(combined_df)
                combined_df = combined_df.drop_duplicates(subset=['PDF File', 'Order Number', 'Line'], keep='last')
                after_dedup = len(combined_df)
                
                duplicates_removed = before_dedup - after_dedup
                if duplicates_removed > 0:
                    update_progress(f"Removed {duplicates_removed} duplicate line items")
                
                update_progress(f"Appended {len(new_df)} new rows, total rows: {after_dedup}")
            else:
                update_progress(f"Creating new file: {output_path}")
                combined_df = new_df
            
            # Write to file
            if output_path.endswith('.xlsx'):
                # Write Excel with proper formatting and auto-fit columns
                self.write_excel_with_formatting(output_path, combined_df)
                update_progress(f"Excel file written successfully: {output_path}")
            else:
                combined_df.to_csv(output_path, index=False, encoding='utf-8')
                update_progress(f"CSV file written successfully: {output_path}")
        
        except PermissionError:
            # Excel file is open - re-raise to be handled by retry logic
            update_progress(f"ERROR: Cannot write to file (file is open or locked)")
            raise
        except Exception as e:
            update_progress(f"Error writing output file: {e}")
            raise
    
    def write_excel_with_formatting(self, output_path, df):
        """Write DataFrame to Excel with formatting and auto-fit columns - EXACT ORIGINAL LOGIC"""
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            
            # Write to Excel using pandas
            df.to_excel(output_path, index=False, sheet_name='PO Data', engine='openpyxl')
            
            # Open workbook to apply formatting
            wb = load_workbook(output_path)
            ws = wb['PO Data']
            
            # Auto-fit column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # Set column width (add a little padding)
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Format Order Number as integer (no decimals, with thousand separators)
            try:
                if 'Order Number' in df.columns:
                    col_idx = list(df.columns).index('Order Number') + 1
                    col_letter = get_column_letter(col_idx)
                    
                    for row in range(2, ws.max_row + 1):
                        cell = ws[f'{col_letter}{row}']
                        if cell.value is not None:
                            cell.number_format = '0'  # Integer with no decimals, no separators
            except:
                pass
            
            # Line column is kept as text (no special formatting needed)
            
            # Format number columns with thousand separators
            from openpyxl.styles import numbers
            numeric_cols = ['Quantity', 'Unit Price', 'Amount']
            
            for col_name in numeric_cols:
                try:
                    # Find column index
                    col_idx = list(df.columns).index(col_name) + 1
                    col_letter = get_column_letter(col_idx)
                    
                    # Apply number format to data rows (skip header)
                    for row in range(2, ws.max_row + 1):
                        cell = ws[f'{col_letter}{row}']
                        if cell.value is not None:
                            if col_name == 'Quantity':
                                # No decimals for quantity
                                cell.number_format = '#,##0'
                            else:
                                # 2 decimals for prices
                                cell.number_format = '#,##0.00'
                except:
                    pass
            
            # Save the formatted workbook
            wb.save(output_path)
            wb.close()
        
        except PermissionError:
            # File is locked or open - re-raise to trigger retry logic
            update_progress(f"ERROR: Cannot write Excel file (file is open or locked)")
            raise
        except Exception as e:
            # If formatting fails, at least we have the basic file
            update_progress(f"Warning: Could not apply Excel formatting: {e}")
            pass

# Create global extractor instance
extractor = PDFExtractor()

def update_progress(message):
    """Update progress and send to UI"""
    print(message)  # Console logging
    eel.update_progress(message)()

@eel.expose
def browse_output_file():
    """Open file dialog for output file selection"""
    root = tk.Tk()
    root.withdraw()
    filename = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")],
        initialfile="PO_Data.xlsx"
    )
    root.destroy()
    return filename if filename else ""

@eel.expose
def load_settings():
    """Load saved settings"""
    try:
        if os.path.exists(extractor.settings_file):
            with open(extractor.settings_file, 'r') as f:
                return json.load(f)
    except:
        pass
    return {}

@eel.expose
def save_settings(settings):
    """Save settings to file"""
    try:
        with open(extractor.settings_file, 'w') as f:
            json.dump(settings, f, indent=2)
        return True
    except:
        return False

@eel.expose
def extract_pdfs_from_outlook(email_addr, folder_text, subject_text, start_date_str, end_date_str, output_path):
    """Main extraction function - runs in background thread WITH EXACT ORIGINAL LOGIC"""
    def run_extraction():
        try:
            # CRITICAL: Initialize COM in this thread
            pythoncom.CoInitialize()
            
            update_progress("Starting PDF extraction...")
            
            # Parse start date - EXACT ORIGINAL LOGIC
            start_date = None
            if start_date_str:
                try:
                    start_date = datetime.strptime(start_date_str, "%m/%d/%Y")
                    update_progress(f"Filter: Emails after {start_date.strftime('%Y-%m-%d')}")
                except ValueError:
                    eel.show_error("Invalid start date format. Use MM/DD/YYYY")()
                    pythoncom.CoUninitialize()
                    return {"success": False, "error": "Invalid start date format. Use MM/DD/YYYY"}
            
            # Parse end date - EXACT ORIGINAL LOGIC
            end_date = None
            if end_date_str:
                try:
                    end_date = datetime.strptime(end_date_str, "%m/%d/%Y")
                    update_progress(f"Filter: Emails before {end_date.strftime('%Y-%m-%d')}")
                except ValueError:
                    eel.show_error("Invalid end date format. Use MM/DD/YYYY")()
                    pythoncom.CoUninitialize()
                    return {"success": False, "error": "Invalid end date format. Use MM/DD/YYYY"}
            
            # Connect to Outlook - CREATE FRESH CONNECTION EACH TIME
            update_progress("Connecting to Outlook...")
            eel.update_status("Connecting to Outlook...")()
            outlook = extractor.connect_outlook()  # This now creates a fresh connection
            
            # Find matching folder - EXACT ORIGINAL LOGIC
            update_progress(f"Searching for folder containing: '{folder_text}'")
            target_folder = extractor.find_folder(outlook, email_addr, folder_text)
            
            if not target_folder:
                eel.show_error(f"Could not find folder containing '{folder_text}' in {email_addr}")()
                pythoncom.CoUninitialize()
                return {"success": False, "error": f"Could not find folder containing '{folder_text}'"}
            
            update_progress(f"Found folder: {target_folder.Name}")
            
            # Filter emails - EXACT ORIGINAL LOGIC
            update_progress(f"Filtering emails with subject containing: '{subject_text}'")
            emails = extractor.filter_emails(target_folder, subject_text, start_date, end_date)
            update_progress(f"Found {len(emails)} matching emails")
            
            if not emails:
                eel.show_warning("No emails match the filter criteria")()
                pythoncom.CoUninitialize()
                return {"success": False, "error": "No emails match the filter criteria"}
            
            # Setup PDF storage location - EXACT ORIGINAL LOGIC
            output_dir = os.path.dirname(output_path)
            pdf_base_folder = os.path.join(output_dir, "PDFs")
            
            # Process PDFs with deduplication - EXACT ORIGINAL LOGIC
            all_data = []
            pdf_count = 0
            processed_pdfs = set()  # Track unique PDFs by hash
            
            for idx, email in enumerate(emails, 1):
                update_progress(f"\n[{idx}/{len(emails)}] Processing: {email.Subject}")
                eel.update_status(f"Processing email {idx}/{len(emails)}...")()
                eel.set_extraction_progress(idx, len(emails))()
                
                # Get email date for folder organization
                email_date = email.ReceivedTime
                date_folder = datetime(email_date.year, email_date.month, email_date.day).strftime("%Y-%m-%d")
                pdf_save_folder = os.path.join(pdf_base_folder, date_folder)
                
                # Create folder if it doesn't exist
                os.makedirs(pdf_save_folder, exist_ok=True)
                
                # Extract PDFs from attachments
                for attachment in email.Attachments:
                    if attachment.FileName.lower().endswith('.pdf'):
                        # Save to temp file first to calculate hash
                        temp_pdf = os.path.join(tempfile.gettempdir(), attachment.FileName)
                        attachment.SaveAsFile(temp_pdf)
                        
                        # Calculate PDF hash for deduplication
                        try:
                            with open(temp_pdf, 'rb') as f:
                                pdf_hash = hashlib.md5(f.read()).hexdigest()
                            
                            # Check if already processed
                            if pdf_hash in processed_pdfs:
                                update_progress(f"  Skipping duplicate: {attachment.FileName}")
                                try:
                                    os.remove(temp_pdf)
                                except:
                                    pass
                                continue
                            
                            # Mark as processed
                            processed_pdfs.add(pdf_hash)
                            pdf_count += 1
                            update_progress(f"  Found PDF: {attachment.FileName}")
                            
                            # Parse PDF
                            data = extractor.parse_pdf(temp_pdf, attachment.FileName)
                            all_data.extend(data)
                            
                            # Save PDF to permanent location with retry on permission error - EXACT ORIGINAL LOGIC
                            permanent_pdf_path = os.path.join(pdf_save_folder, attachment.FileName)
                            for pdf_attempt in range(3):  # Try up to 3 times for PDF saves
                                try:
                                    shutil.copy2(temp_pdf, permanent_pdf_path)
                                    update_progress(f"  Saved to: {date_folder}/{attachment.FileName}")
                                    break
                                except PermissionError:
                                    if pdf_attempt < 2:
                                        # Send retry request to UI
                                        result = eel.ask_retry_pdf(attachment.FileName)()
                                        if not result:
                                            update_progress(f"  User cancelled PDF save for: {attachment.FileName}")
                                            break
                                    else:
                                        update_progress(f"  ERROR: Could not save PDF after 3 attempts: {attachment.FileName}")
                                        break
                                except Exception as e:
                                    update_progress(f"  Warning: Could not save PDF: {e}")
                                    break
                        
                        except Exception as e:
                            update_progress(f"  Error processing {attachment.FileName}: {e}")
                        
                        # Clean up temp file
                        try:
                            os.remove(temp_pdf)
                        except:
                            pass
            
            # Write to file (Excel or CSV) with retry on permission error - EXACT ORIGINAL LOGIC
            update_progress(f"\nWriting {len(all_data)} line items to file...")
            max_retries = 5
            for attempt in range(max_retries):
                try:
                    extractor.write_output(output_path, all_data)
                    break  # Success, exit retry loop
                except PermissionError:
                    if attempt < max_retries - 1:
                        # Ask user to retry
                        result = eel.ask_retry_file(output_path, attempt + 2, max_retries)()
                        if not result:
                            update_progress("User cancelled file write operation")
                            eel.show_info("Data extraction completed but file was not saved.")()
                            pythoncom.CoUninitialize()
                            return {"success": False, "error": "User cancelled file write"}
                        else:
                            update_progress(f"Retrying file write (attempt {attempt + 2}/{max_retries})...")
                    else:
                        # Max retries reached
                        update_progress("ERROR: Max retries reached, file could not be written")
                        eel.show_error("Could not write file after multiple attempts.")()
                        pythoncom.CoUninitialize()
                        return {"success": False, "error": "Could not write file after multiple attempts"}
            
            update_progress(f"\n{'='*60}")
            update_progress(f"SUCCESS! Extracted {pdf_count} PDFs with {len(all_data)} line items")
            update_progress(f"Output saved to: {output_path}")
            update_progress(f"PDFs saved to: {pdf_base_folder}")
            update_progress(f"{'='*60}")
            
            eel.update_status(f"Complete! {len(all_data)} items extracted")()
            
            # Ask to open file - EXACT ORIGINAL LOGIC
            eel.extraction_complete_with_prompt(len(all_data), output_path)()
            
            # Clean up COM
            pythoncom.CoUninitialize()
            
            return {"success": True, "items": len(all_data), "pdfs": pdf_count}
        
        except Exception as e:
            error_msg = f"ERROR: {str(e)}"
            update_progress(error_msg)
            eel.show_error(f"An error occurred:\n{str(e)}")()
            eel.update_status("Error occurred")()
            
            # Clean up COM on error
            try:
                pythoncom.CoUninitialize()
            except:
                pass
            
            return {"success": False, "error": str(e)}
    
    # Run in background thread
    thread = threading.Thread(target=run_extraction)
    thread.daemon = True
    thread.start()
    return {"started": True}

@eel.expose
def open_file(filepath):
    """Open file with default application"""
    try:
        normalized_path = os.path.abspath(os.path.normpath(filepath))
        os.startfile(normalized_path)
        return True
    except Exception as e:
        print(f"Error opening file: {e}")
        return False

# Start the application
if __name__ == "__main__":
    eel.start('index.html', 
        mode="edge", 
        size=(850, 750),
        port=0, 
        app_mode=True, 
        disable_cache=True
    )